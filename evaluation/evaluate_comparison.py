"""
evaluation/evaluate_comparison.py
────────────────────────────────────────────────────────────
Model Comparison Evaluator: Baseline vs Edu-VQAGuider

Compares:
  BASELINE : Qwen2.5-3B answering WITHOUT any video context (zero-shot)
  EDU-VQAGUIDER: Full RAG pipeline (Whisper → Chunk → Retrieve → Qwen)

Metrics computed per answer:
  - ROUGE-L F1        (lexical overlap with reference)
  - BLEU-4            (n-gram precision)
  - Answer Length     (word count)
  - Response Time     (seconds)
  - Route (EDU only)  (concept/procedure/temporal/visual/summary)
  - Confidence        (planner confidence %)

Output:
  evaluation/results/comparison_results.xlsx   (colored Excel file)
  evaluation/results/comparison_results.csv    (raw data)

Usage:
  # Make sure your FastAPI backend is running first:
  #   uvicorn backend.main:app --host 0.0.0.0 --port 8000

  python evaluation/evaluate_comparison.py \\
      --csv evaluation/test_questions.csv \\
      --video_path path/to/your/lecture.mp4 \\
      --backend http://localhost:8000

  # To use a pre-uploaded video (already processed):
  python evaluation/evaluate_comparison.py \\
      --csv evaluation/test_questions.csv \\
      --video_id <your-video-uuid> \\
      --backend http://localhost:8000
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Optional

import requests

# ── Try importing optional scoring libraries ─────────────────
try:
    from rouge_score import rouge_scorer
    ROUGE_AVAILABLE = True
except ImportError:
    ROUGE_AVAILABLE = False
    print("WARNING: rouge-score not installed. Install with: pip install rouge-score")

try:
    from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
    import nltk
    nltk.download("punkt", quiet=True)
    BLEU_AVAILABLE = True
except ImportError:
    BLEU_AVAILABLE = False
    print("WARNING: nltk not installed. Install with: pip install nltk")

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Install with: pip install openpyxl")

# ─────────────────────────────────────────────────────────────
# Baseline Model: Qwen direct (no context)
# ─────────────────────────────────────────────────────────────

def baseline_answer(question: str, backend_url: str) -> tuple[str, float]:
    """
    Get answer from BASELINE model: Qwen2.5-3B with NO video context.
    
    Calls a special /api/v2/baseline_ask endpoint that skips RAG.
    If that endpoint doesn't exist, falls back to direct Qwen call.
    
    Returns:
        (answer_text, response_time_seconds)
    """
    t0 = time.time()
    
    # Try the baseline endpoint first
    try:
        resp = requests.post(
            f"{backend_url.rstrip('/')}/api/v2/baseline_ask",
            json={"question": question},
            timeout=120,
        )
        if resp.ok:
            data = resp.json()
            return data.get("answer", ""), time.time() - t0
    except Exception:
        pass
    
    # Fallback: call Qwen directly via a simple prompt (no context)
    # This simulates "asking the LLM without watching the video"
    baseline_prompt = (
        f"You are an educational AI assistant. Answer the following question "
        f"based on your general knowledge. Do not assume you have seen any specific video.\n\n"
        f"Question: {question}\n\nAnswer:"
    )
    
    try:
        resp = requests.post(
            f"{backend_url.rstrip('/')}/api/v2/direct_generate",
            json={"prompt": baseline_prompt},
            timeout=120,
        )
        if resp.ok:
            data = resp.json()
            return data.get("answer", ""), time.time() - t0
    except Exception:
        pass
    
    # Ultimate fallback: return placeholder
    elapsed = time.time() - t0
    return "[Baseline endpoint not available - add /api/v2/baseline_ask to backend]", elapsed


def edu_vqaguider_answer(
    video_id: str,
    question: str,
    backend_url: str,
) -> tuple[str, str, float, float, str, str]:
    """
    Get answer from Edu-VQAGuider (full RAG pipeline).
    
    Returns:
        (direct_answer, detailed_answer, confidence, response_time,
         route, planner_source)
    """
    t0 = time.time()
    url = f"{backend_url.rstrip('/')}/api/v2/videos/{video_id}/ask"
    
    try:
        resp = requests.post(
            url,
            json={"question": question},
            timeout=300,
        )
        resp.raise_for_status()
        data = resp.json()
        
        route_info = data.get("route", {})
        elapsed = time.time() - t0
        
        return (
            data.get("direct_answer", ""),
            data.get("detailed_answer", ""),
            route_info.get("confidence", 0.0),
            elapsed,
            route_info.get("route", "unknown"),
            route_info.get("planner_source", "unknown"),
        )
    except requests.exceptions.ConnectionError:
        return ("ERROR: Cannot connect to backend", "", 0.0, 0.0, "error", "error")
    except Exception as e:
        return (f"ERROR: {e}", "", 0.0, 0.0, "error", "error")


# ─────────────────────────────────────────────────────────────
# Metrics
# ─────────────────────────────────────────────────────────────

def compute_rouge_l(hypothesis: str, reference: str) -> float:
    """Compute ROUGE-L F1 score."""
    if not ROUGE_AVAILABLE or not hypothesis or not reference:
        return 0.0
    scorer = rouge_scorer.RougeScorer(["rougeL"], use_stemmer=True)
    scores = scorer.score(reference, hypothesis)
    return round(scores["rougeL"].fmeasure, 4)


def compute_bleu4(hypothesis: str, reference: str) -> float:
    """Compute BLEU-4 score."""
    if not BLEU_AVAILABLE or not hypothesis or not reference:
        return 0.0
    try:
        ref_tokens = reference.lower().split()
        hyp_tokens = hypothesis.lower().split()
        if not hyp_tokens:
            return 0.0
        smooth = SmoothingFunction().method1
        score = sentence_bleu(
            [ref_tokens], hyp_tokens,
            weights=(0.25, 0.25, 0.25, 0.25),
            smoothing_function=smooth,
        )
        return round(score, 4)
    except Exception:
        return 0.0


def word_count(text: str) -> int:
    """Count words in a string."""
    return len(text.split()) if text else 0


def compute_winner(baseline_rouge: float, edu_rouge: float) -> str:
    """Determine winner based on ROUGE-L scores."""
    diff = edu_rouge - baseline_rouge
    if diff > 0.05:
        return "Edu-VQAGuider ✓"
    elif diff < -0.05:
        return "Baseline ✓"
    else:
        return "Tie"


# ─────────────────────────────────────────────────────────────
# Video Setup
# ─────────────────────────────────────────────────────────────

def upload_and_transcribe(video_path: str, backend_url: str) -> Optional[str]:
    """Upload a video and transcribe it. Returns video_id or None."""
    print(f"\n📤 Uploading video: {video_path}")
    video_path = Path(video_path)
    
    if not video_path.exists():
        print(f"ERROR: Video file not found: {video_path}")
        return None
    
    # Upload
    with open(video_path, "rb") as f:
        try:
            resp = requests.post(
                f"{backend_url.rstrip('/')}/api/v2/videos",
                files={"file": (video_path.name, f, "video/mp4")},
                timeout=120,
            )
            resp.raise_for_status()
            data = resp.json()
            video_id = data["video_id"]
            print(f"✅ Uploaded. video_id={video_id} | chunks={data.get('num_chunks', '?')}")
        except Exception as e:
            print(f"ERROR uploading video: {e}")
            return None
    
    # Transcribe
    print("🎙️ Starting Whisper transcription (this may take a few minutes)...")
    try:
        resp = requests.post(
            f"{backend_url.rstrip('/')}/api/v2/videos/{video_id}/transcribe",
            timeout=600,
        )
        resp.raise_for_status()
        data = resp.json()
        print(f"✅ Transcribed. Chunks with text: {data.get('num_chunks_with_text', '?')}")
    except Exception as e:
        print(f"ERROR transcribing: {e}")
        return None
    
    return video_id


# ─────────────────────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────────────────────

ROUTE_COLORS = {
    "concept":   "DBEAFE",
    "procedure": "DCFCE7",
    "temporal":  "FEF3C7",
    "visual":    "EDE9FE",
    "summary":   "FCE7F3",
    "unknown":   "F1F5F9",
    "error":     "FEE2E2",
}

WINNER_COLORS = {
    "Edu-VQAGuider ✓": "D1FAE5",  # green
    "Baseline ✓":       "FEE2E2",  # red
    "Tie":              "FEF9C3",  # yellow
}


def export_excel(rows: list[dict], output_path: str):
    """Export comparison results to a styled Excel file."""
    if not EXCEL_AVAILABLE:
        print("Excel export skipped (openpyxl not installed)")
        return
    
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Main Comparison ──────────────────────────────
    ws = wb.active
    ws.title = "Comparison"
    
    headers = [
        "No.",
        "Category",
        "Question",
        "Reference Answer",
        "── BASELINE ──",
        "Baseline Answer",
        "Baseline ROUGE-L",
        "Baseline BLEU-4",
        "Baseline Words",
        "Baseline Time (s)",
        "── EDU-VQAGUIDER ──",
        "EDU Direct Answer",
        "EDU Detailed Answer",
        "EDU ROUGE-L",
        "EDU BLEU-4",
        "EDU Words",
        "EDU Time (s)",
        "EDU Route",
        "EDU Confidence",
        "EDU Planner Source",
        "── VERDICT ──",
        "Winner (ROUGE-L)",
        "ROUGE-L Improvement",
    ]
    
    # Header style
    header_fill_main = PatternFill("solid", fgColor="1E3A5F")
    header_fill_base = PatternFill("solid", fgColor="6B21A8")
    header_fill_edu  = PatternFill("solid", fgColor="065F46")
    header_fill_verdict = PatternFill("solid", fgColor="78350F")
    
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        
        # Color code header sections
        if col_idx <= 4:
            cell.fill = header_fill_main
        elif "BASELINE" in header or (5 <= col_idx <= 10):
            cell.fill = header_fill_base
        elif "EDU" in header or (10 <= col_idx <= 20):
            cell.fill = header_fill_edu
        else:
            cell.fill = header_fill_verdict
    
    ws.row_dimensions[1].height = 40
    
    # Data rows
    for row_idx, row in enumerate(rows, 2):
        data = [
            row["no"],
            row["category"],
            row["question"],
            row["reference_answer"],
            "",  # separator
            row["baseline_answer"][:500] + "..." if len(row["baseline_answer"]) > 500 else row["baseline_answer"],
            row["baseline_rouge_l"],
            row["baseline_bleu4"],
            row["baseline_words"],
            round(row["baseline_time"], 2),
            "",  # separator
            row["edu_direct_answer"],
            row["edu_detailed_answer"][:500] + "..." if len(row["edu_detailed_answer"]) > 500 else row["edu_detailed_answer"],
            row["edu_rouge_l"],
            row["edu_bleu4"],
            row["edu_words"],
            round(row["edu_time"], 2),
            row["edu_route"],
            f"{row['edu_confidence']:.1%}",
            row["edu_planner_source"],
            "",  # separator
            row["winner"],
            f"{row['rouge_improvement']:+.4f}",
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            
            # Alternating row background
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
        
        # Color route cell
        route = row["edu_route"]
        route_fill_color = ROUTE_COLORS.get(route, "F1F5F9")
        ws.cell(row=row_idx, column=18).fill = PatternFill("solid", fgColor=route_fill_color)
        ws.cell(row=row_idx, column=18).font = Font(bold=True)
        
        # Color winner cell
        winner = row["winner"]
        winner_color = WINNER_COLORS.get(winner, "F1F5F9")
        ws.cell(row=row_idx, column=22).fill = PatternFill("solid", fgColor=winner_color)
        ws.cell(row=row_idx, column=22).font = Font(bold=True)
        
        # Color improvement cell (green if positive, red if negative)
        imp = row["rouge_improvement"]
        if imp > 0.05:
            ws.cell(row=row_idx, column=23).fill = PatternFill("solid", fgColor="D1FAE5")
            ws.cell(row=row_idx, column=23).font = Font(bold=True, color="065F46")
        elif imp < -0.05:
            ws.cell(row=row_idx, column=23).fill = PatternFill("solid", fgColor="FEE2E2")
            ws.cell(row=row_idx, column=23).font = Font(bold=True, color="991B1B")
    
    # Column widths
    col_widths = {
        1: 5, 2: 12, 3: 40, 4: 35, 5: 16, 6: 45, 7: 14, 8: 12,
        9: 12, 10: 14, 11: 18, 12: 40, 13: 55, 14: 12, 15: 12,
        16: 12, 17: 12, 18: 14, 19: 14, 20: 16, 21: 14, 22: 18, 23: 18,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "C2"
    
    # ── Sheet 2: Summary Statistics ───────────────────────────
    ws2 = wb.create_sheet("Summary")
    
    # Compute aggregate stats
    n = len(rows)
    if n == 0:
        wb.save(output_path)
        return
    
    baseline_rouges = [r["baseline_rouge_l"] for r in rows]
    edu_rouges      = [r["edu_rouge_l"] for r in rows]
    baseline_bleus  = [r["baseline_bleu4"] for r in rows]
    edu_bleus       = [r["edu_bleu4"] for r in rows]
    baseline_times  = [r["baseline_time"] for r in rows]
    edu_times       = [r["edu_time"] for r in rows]
    baseline_words  = [r["baseline_words"] for r in rows]
    edu_words       = [r["edu_words"] for r in rows]
    
    edu_wins = sum(1 for r in rows if "Edu" in r["winner"])
    base_wins = sum(1 for r in rows if "Baseline" in r["winner"])
    ties = n - edu_wins - base_wins
    
    # Route distribution
    route_counts: dict[str, int] = {}
    for r in rows:
        route_counts[r["edu_route"]] = route_counts.get(r["edu_route"], 0) + 1
    
    summary_data = [
        ["Edu-VQAGuider vs Baseline — Evaluation Summary", ""],
        ["", ""],
        ["METRIC", "BASELINE", "EDU-VQAGUIDER", "IMPROVEMENT"],
        ["Avg ROUGE-L",    f"{sum(baseline_rouges)/n:.4f}", f"{sum(edu_rouges)/n:.4f}",    f"{(sum(edu_rouges)-sum(baseline_rouges))/n:+.4f}"],
        ["Avg BLEU-4",     f"{sum(baseline_bleus)/n:.4f}",  f"{sum(edu_bleus)/n:.4f}",     f"{(sum(edu_bleus)-sum(baseline_bleus))/n:+.4f}"],
        ["Avg Answer Length (words)", f"{sum(baseline_words)/n:.1f}", f"{sum(edu_words)/n:.1f}", ""],
        ["Avg Response Time (s)",     f"{sum(baseline_times)/n:.2f}", f"{sum(edu_times)/n:.2f}", ""],
        ["Total Questions", str(n), str(n), ""],
        ["", ""],
        ["WINS", "Count", "Percentage", ""],
        ["Edu-VQAGuider wins",  str(edu_wins),  f"{edu_wins/n:.1%}",  ""],
        ["Baseline wins",       str(base_wins), f"{base_wins/n:.1%}", ""],
        ["Ties",                str(ties),       f"{ties/n:.1%}",      ""],
        ["", ""],
        ["ROUTE DISTRIBUTION (Edu-VQAGuider)", "", "", ""],
    ]
    
    for route, count in sorted(route_counts.items(), key=lambda x: -x[1]):
        summary_data.append([f"  {route}", str(count), f"{count/n:.1%}", ""])
    
    # Write summary
    title_fill = PatternFill("solid", fgColor="1E3A5F")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header2_fill = PatternFill("solid", fgColor="374151")
    header2_font = Font(bold=True, color="FFFFFF")
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Style title row
        if row_idx == 1:
            for col_idx in range(1, 5):
                ws2.cell(row=row_idx, column=col_idx).fill = title_fill
                ws2.cell(row=row_idx, column=col_idx).font = title_font
            ws2.merge_cells(f"A1:D1")
        
        # Style header rows
        if row_data and row_data[0] in ["METRIC", "WINS", "ROUTE DISTRIBUTION (Edu-VQAGuider)"]:
            for col_idx in range(1, 5):
                ws2.cell(row=row_idx, column=col_idx).fill = header2_fill
                ws2.cell(row=row_idx, column=col_idx).font = header2_font
        
        # Color edu-wins green
        if row_data and "Edu-VQAGuider wins" in str(row_data[0]):
            for col_idx in range(1, 5):
                ws2.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="D1FAE5")
                ws2.cell(row=row_idx, column=col_idx).font = Font(bold=True, color="065F46")
        
        # Color baseline-wins red
        if row_data and "Baseline wins" in str(row_data[0]):
            for col_idx in range(1, 5):
                ws2.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="FEE2E2")
                ws2.cell(row=row_idx, column=col_idx).font = Font(bold=True, color="991B1B")
    
    for col_idx, width in {1: 35, 2: 20, 3: 20, 4: 20}.items():
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws2.row_dimensions[1].height = 30
    
    # ── Sheet 3: Per-Category Analysis ────────────────────────
    ws3 = wb.create_sheet("By Category")
    
    categories = sorted(set(r["category"] for r in rows))
    
    cat_headers = ["Category", "Count",
                   "Baseline Avg ROUGE-L", "EDU Avg ROUGE-L",
                   "ROUGE-L Gain", "EDU Wins", "Baseline Wins"]
    
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    for row_idx, cat in enumerate(categories, 2):
        cat_rows = [r for r in rows if r["category"] == cat]
        n_cat = len(cat_rows)
        avg_base = sum(r["baseline_rouge_l"] for r in cat_rows) / n_cat
        avg_edu  = sum(r["edu_rouge_l"] for r in cat_rows) / n_cat
        gain = avg_edu - avg_base
        cat_edu_wins  = sum(1 for r in cat_rows if "Edu" in r["winner"])
        cat_base_wins = sum(1 for r in cat_rows if "Baseline" in r["winner"])
        
        row_data = [cat, n_cat, f"{avg_base:.4f}", f"{avg_edu:.4f}",
                    f"{gain:+.4f}", cat_edu_wins, cat_base_wins]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Color route cell
        route_color = ROUTE_COLORS.get(cat, "F1F5F9")
        ws3.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=route_color)
        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Color gain cell
        gain_cell = ws3.cell(row=row_idx, column=5)
        if gain > 0.02:
            gain_cell.fill = PatternFill("solid", fgColor="D1FAE5")
            gain_cell.font = Font(bold=True, color="065F46")
        elif gain < -0.02:
            gain_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            gain_cell.font = Font(bold=True, color="991B1B")
    
    for col_idx, width in {1: 16, 2: 10, 3: 22, 4: 22, 5: 18, 6: 12, 7: 14}.items():
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Excel saved: {output_path}")


# ─────────────────────────────────────────────────────────────
# CSV Export
# ─────────────────────────────────────────────────────────────

def export_csv(rows: list[dict], output_path: str):
    """Export comparison results to CSV."""
    if not rows:
        return
    
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅ CSV saved: {output_path}")


# ─────────────────────────────────────────────────────────────
# Main Evaluation Loop
# ─────────────────────────────────────────────────────────────

def run_evaluation(
    csv_path: str,
    video_id: str,
    backend_url: str,
    output_dir: str = "evaluation/results",
) -> list[dict]:
    """
    Run the full evaluation loop.
    
    Reads questions from CSV, queries both models, computes metrics.
    """
    # Read test questions
    questions = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            questions.append(row)
    
    print(f"\n📋 Loaded {len(questions)} test questions from {csv_path}")
    print(f"🎯 Video ID: {video_id}")
    print(f"🌐 Backend: {backend_url}")
    print("─" * 70)
    
    rows = []
    
    for i, q in enumerate(questions, 1):
        question = q.get("question", "").strip()
        reference = q.get("reference_answer", "").strip()
        category  = q.get("category", "unknown").strip()
        video_url = q.get("video_url", "").strip()
        
        if not question:
            continue
        
        print(f"\n[{i}/{len(questions)}] {category.upper()} — {question[:60]}...")
        
        # ── Baseline ─────────────────────────────────────────
        print("  → Querying BASELINE...")
        base_answer, base_time = baseline_answer(question, backend_url)
        base_rouge = compute_rouge_l(base_answer, reference)
        base_bleu  = compute_bleu4(base_answer, reference)
        base_words = word_count(base_answer)
        print(f"     ROUGE-L={base_rouge:.4f} | BLEU-4={base_bleu:.4f} | {base_time:.1f}s")
        
        # ── Edu-VQAGuider ─────────────────────────────────────
        print("  → Querying EDU-VQAGUIDER...")
        (edu_direct, edu_detailed, edu_conf, edu_time,
         edu_route, edu_source) = edu_vqaguider_answer(video_id, question, backend_url)
        
        # Use direct answer for metric comparison (concise, comparable)
        edu_answer = edu_direct if edu_direct else edu_detailed
        edu_rouge  = compute_rouge_l(edu_answer, reference)
        edu_bleu   = compute_bleu4(edu_answer, reference)
        edu_words  = word_count(edu_detailed)  # count detailed for richness
        print(f"     Route={edu_route} | Conf={edu_conf:.1%} | "
              f"ROUGE-L={edu_rouge:.4f} | BLEU-4={edu_bleu:.4f} | {edu_time:.1f}s")
        
        # ── Winner ───────────────────────────────────────────
        winner = compute_winner(base_rouge, edu_rouge)
        improvement = edu_rouge - base_rouge
        print(f"  → WINNER: {winner} (improvement: {improvement:+.4f})")
        
        rows.append({
            "no": i,
            "video_url": video_url,
            "category": category,
            "question": question,
            "reference_answer": reference,
            "baseline_answer": base_answer,
            "baseline_rouge_l": base_rouge,
            "baseline_bleu4": base_bleu,
            "baseline_words": base_words,
            "baseline_time": base_time,
            "edu_direct_answer": edu_direct,
            "edu_detailed_answer": edu_detailed,
            "edu_rouge_l": edu_rouge,
            "edu_bleu4": edu_bleu,
            "edu_words": edu_words,
            "edu_time": edu_time,
            "edu_route": edu_route,
            "edu_confidence": edu_conf,
            "edu_planner_source": edu_source,
            "winner": winner,
            "rouge_improvement": improvement,
        })
    
    # Summary
    n = len(rows)
    if n > 0:
        edu_wins  = sum(1 for r in rows if "Edu" in r["winner"])
        base_wins = sum(1 for r in rows if "Baseline" in r["winner"])
        avg_base_rouge = sum(r["baseline_rouge_l"] for r in rows) / n
        avg_edu_rouge  = sum(r["edu_rouge_l"] for r in rows) / n
        
        print("\n" + "═" * 70)
        print("  FINAL RESULTS")
        print("═" * 70)
        print(f"  Total questions:       {n}")
        print(f"  Edu-VQAGuider wins:   {edu_wins}/{n} ({edu_wins/n:.1%})")
        print(f"  Baseline wins:        {base_wins}/{n} ({base_wins/n:.1%})")
        print(f"  Avg ROUGE-L Baseline: {avg_base_rouge:.4f}")
        print(f"  Avg ROUGE-L Edu:      {avg_edu_rouge:.4f}")
        print(f"  Overall improvement:  {avg_edu_rouge - avg_base_rouge:+.4f}")
        print("═" * 70)
    
    return rows


# ─────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Compare Baseline vs Edu-VQAGuider on educational video QA"
    )
    parser.add_argument(
        "--csv", type=str, required=True,
        help="Path to test questions CSV (see evaluation/test_questions.csv)"
    )
    parser.add_argument(
        "--video_path", type=str, default=None,
        help="Path to video file (will upload + transcribe automatically)"
    )
    parser.add_argument(
        "--video_id", type=str, default=None,
        help="Existing video_id (if already uploaded and transcribed)"
    )
    parser.add_argument(
        "--backend", type=str, default="http://localhost:8000",
        help="Backend URL (default: http://localhost:8000)"
    )
    parser.add_argument(
        "--output_dir", type=str, default="evaluation/results",
        help="Directory for output files"
    )
    args = parser.parse_args()
    
    # Check backend
    try:
        resp = requests.get(f"{args.backend}/health", timeout=5)
        data = resp.json()
        print(f"✅ Backend online | v2_edu_ready={data.get('v2_edu_ready')} | "
              f"device={data.get('device')}")
    except Exception:
        print(f"❌ Cannot reach backend at {args.backend}")
        print("   Start it with: uvicorn backend.main:app --host 0.0.0.0 --port 8000")
        sys.exit(1)
    
    # Get or upload video
    video_id = args.video_id
    if video_id is None:
        if args.video_path is None:
            print("ERROR: Provide either --video_id or --video_path")
            sys.exit(1)
        video_id = upload_and_transcribe(args.video_path, args.backend)
        if video_id is None:
            print("ERROR: Video upload/transcription failed")
            sys.exit(1)
    
    # Run evaluation
    rows = run_evaluation(
        csv_path=args.csv,
        video_id=video_id,
        backend_url=args.backend,
        output_dir=args.output_dir,
    )
    
    # Export
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    excel_path = str(Path(args.output_dir) / f"comparison_{timestamp}.xlsx")
    csv_out    = str(Path(args.output_dir) / f"comparison_{timestamp}.csv")
    
    export_excel(rows, excel_path)
    export_csv(rows, csv_out)
    
    print(f"\n📊 Results saved to: {args.output_dir}/")


if __name__ == "__main__":
    main()
